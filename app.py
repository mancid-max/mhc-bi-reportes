from __future__ import annotations

import csv
import os
from pathlib import Path
from flask import Flask, render_template

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import openpyxl as _openpyxl
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

try:
    import requests as _requests
    _REQUESTS = True
except ImportError:
    _REQUESTS = False

app = Flask(__name__)

BI_DIR   = Path(os.environ.get("BI_REPORTES_BI_DIR", r"Z:\BI"))
SEED_DIR = Path(__file__).parent / "seed"

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

COLECCIONES           = ["40", "41", "42", "43"]
TEMPORADAS_HISTORICAS = {"40", "41", "42"}
TEMPORADAS_ACTUALES   = {"43", "44"}


def _read_csv(path: Path) -> list[dict]:
    for enc in ("utf-8-sig", "cp1252", "latin-1", "cp1250"):
        try:
            with open(path, encoding=enc, errors="strict", newline="") as f:
                return list(csv.DictReader(f, delimiter=";"))
        except (UnicodeDecodeError, Exception):
            continue
    # fallback con reemplazo
    try:
        with open(path, encoding="cp1252", errors="replace", newline="") as f:
            return list(csv.DictReader(f, delimiter=";"))
    except Exception:
        return []


def _to_int(v) -> int:
    try:
        return int(str(v or "").strip().replace(".", "").replace(",", ""))
    except Exception:
        return 0


def _ventas_path() -> Path:
    p = BI_DIR / "VENTAS-TOD-2026.CSV"
    return p if p.exists() and p.stat().st_size > 0 else SEED_DIR / "VENTAS-TOD-2026.CSV"


def _fetch_quotes() -> list[dict]:
    if not _REQUESTS:
        return []
    try:
        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
        }
        # Traer todos los quotes
        r = _requests.get(
            f"{SUPABASE_URL}/rest/v1/quotes",
            headers=headers,
            params={"select": "id,store_name,client_rut_normalized,total_items,created_at,is_ready,source"},
            timeout=8,
        )
        quotes = r.json() if r.ok else []

        # Traer todos los quote_items
        r2 = _requests.get(
            f"{SUPABASE_URL}/rest/v1/quote_items",
            headers=headers,
            params={"select": "quote_id,sku,size,quantity"},
            timeout=8,
        )
        items = r2.json() if r2.ok else []

        # Agrupar items por quote_id
        items_by_quote: dict[str, list] = {}
        for item in items:
            qid = item["quote_id"]
            items_by_quote.setdefault(qid, []).append(item)

        for q in quotes:
            q["items"] = items_by_quote.get(q["id"], [])

        return quotes
    except Exception:
        return []


def _numeros_path() -> Path:
    p = SEED_DIR / "Numeros.xlsx"
    return p if p.exists() else Path()


def _data_personal_path() -> Path:
    p = SEED_DIR / "Data personal clientes.TXT"
    return p if p.exists() else Path()


def _pedidos_path() -> Path:
    p = SEED_DIR / "PEDIDOS FOR BI.Txt"
    return p if p.exists() and p.stat().st_size > 0 else Path()


# ── Reporte 1: Mejores clientes (todas las colecciones) ──────────────────────

def report_mejores_clientes() -> dict:
    path  = _ventas_path()
    empty = {"available": False, "rows": [], "total_ventas": 0, "total_prendas": 0, "total_documentos": 0}
    if not path.exists():
        return {c: empty for c in COLECCIONES}

    rows = _read_csv(path)
    data: dict[str, dict] = {c: {} for c in COLECCIONES}

    for r in rows:
        temp = str(r.get("Temporada") or "").strip()
        if temp not in COLECCIONES:
            continue
        rut    = str(r.get("Rut")     or "").strip()
        nombre = str(r.get("cliente") or "").strip() or rut
        total  = _to_int(r.get("Total"))
        cant   = _to_int(r.get("Cant"))
        num    = str(r.get("Numero")  or "").strip()
        key    = rut or nombre

        if key not in data[temp]:
            data[temp][key] = {"rut": rut, "cliente": nombre,
                               "total": 0, "prendas": 0, "documentos": set()}
        data[temp][key]["total"]    += total
        data[temp][key]["prendas"]  += cant
        data[temp][key]["documentos"].add(num)

    result = {}
    for cole in COLECCIONES:
        filas = sorted(
            [
                {
                    "rut":       v["rut"],
                    "cliente":   v["cliente"],
                    "total":     v["total"],
                    "prendas":   v["prendas"],
                    "documentos": len(v["documentos"]),
                }
                for v in data[cole].values() if v["total"] > 0 and v["prendas"] > 5
            ],
            key=lambda x: -x["total"],
        )
        result[cole] = {
            "available":        bool(filas),
            "rows":             filas,
            "total_ventas":     sum(r["total"]      for r in filas),
            "total_prendas":    sum(r["prendas"]    for r in filas),
            "total_documentos": sum(r["documentos"] for r in filas),
        }
    return result


# ── Reporte 2: Clientes que no compraron ─────────────────────────────────────

def report_no_compraron() -> dict:
    path = _ventas_path()
    if not path.exists():
        return {"available": False, "rows": [], "total": 0}

    rows = _read_csv(path)
    # Clientes que compraron en C42 con más de 10 prendas
    en_c42:  dict[str, dict] = {}
    # Clientes que compraron en C43
    en_c43:  set[str]        = set()

    for r in rows:
        temp = str(r.get("Temporada") or "").strip()
        rut    = str(r.get("Rut")     or "").strip()
        nombre = str(r.get("cliente") or "").strip() or rut
        key    = rut or nombre

        if temp == "42":
            if key not in en_c42:
                en_c42[key] = {"rut": rut, "cliente": nombre, "total": 0, "prendas": 0}
            en_c42[key]["total"]   += _to_int(r.get("Total"))
            en_c42[key]["prendas"] += _to_int(r.get("Cant"))

        if temp == "43":
            en_c43.add(key)

    no_compraron = sorted(
        [
            {
                "rut":             v["rut"],
                "cliente":         v["cliente"],
                "total_historico": v["total"],
                "prendas_c42":     v["prendas"],
            }
            for key, v in en_c42.items()
            if key not in en_c43 and v["prendas"] > 5
        ],
        key=lambda x: -x["total_historico"],
    )
    return {"available": True, "rows": no_compraron, "total": len(no_compraron)}


# ── Reporte 3: Modelo + Pedidos por Bota (todas las colecciones) ──────────────

def report_modelo_bota() -> dict:
    path  = _pedidos_path()
    empty = {"available": False, "botas": [], "modelos": [],
             "total_solicitado": 0, "total_despachado": 0, "total_saldo": 0}
    if not path.exists():
        return {c: empty for c in COLECCIONES}

    rows = _read_csv(path)
    cole_data: dict[str, dict] = {c: {"botas": {}, "modelos": {}} for c in COLECCIONES}

    for r in rows:
        cole = str(r.get("COLECCION") or "").strip()
        if cole not in COLECCIONES:
            continue

        bota       = str(r.get("SubCateg")    or "").strip() or "Sin categoría"
        categ      = str(r.get("Categ")       or "").strip()
        art        = str(r.get("ARTICULO")    or "").strip()
        desc       = str(r.get("DESCRIPCION") or "").strip()
        solicitado = _to_int(r.get("SOLICITADO"))
        despachado = _to_int(r.get("DESPACHADO"))
        saldo      = _to_int(r.get("saldo"))

        botas = cole_data[cole]["botas"]
        if bota not in botas:
            botas[bota] = {"bota": bota, "categ": categ,
                           "solicitado": 0, "despachado": 0, "saldo": 0, "clientes": set()}
        botas[bota]["solicitado"] += solicitado
        botas[bota]["despachado"] += despachado
        botas[bota]["saldo"]      += saldo
        botas[bota]["clientes"].add(str(r.get("RUT") or "").strip())

        modelos = cole_data[cole]["modelos"]
        if art not in modelos:
            modelos[art] = {"articulo": art, "descripcion": desc, "bota": bota,
                            "solicitado": 0, "despachado": 0, "saldo": 0}
        modelos[art]["solicitado"] += solicitado
        modelos[art]["despachado"] += despachado
        modelos[art]["saldo"]      += saldo

    result = {}
    for cole in COLECCIONES:
        botas_raw   = cole_data[cole]["botas"]
        modelos_raw = cole_data[cole]["modelos"]
        botas_list  = sorted(
            [
                {**v, "clientes": len(v["clientes"]),
                 "pct_despachado": round(v["despachado"] / v["solicitado"] * 100) if v["solicitado"] else 0}
                for v in botas_raw.values()
            ],
            key=lambda x: -x["solicitado"],
        )
        modelos_list = sorted(modelos_raw.values(), key=lambda x: -x["solicitado"])
        ts = sum(b["solicitado"] for b in botas_list)
        td = sum(b["despachado"] for b in botas_list)
        result[cole] = {
            "available":        bool(botas_list),
            "botas":            botas_list,
            "modelos":          modelos_list,
            "total_solicitado": ts,
            "total_despachado": td,
            "total_saldo":      sum(b["saldo"] for b in botas_list),
        }
    return result


# ── Reporte 4: Ficha de clientes (C40 en adelante) ───────────────────────────

def report_ficha_clientes() -> list:
    path = _ventas_path()
    if not path.exists():
        return []

    def _norm_rut(s: str) -> str:
        return str(s or "").strip().replace(".", "").replace(",", "").replace("-", "")

    # Cargar mail desde Data personal clientes.TXT
    personal_mail: dict[str, str] = {}
    dp = _data_personal_path()
    if dp.exists():
        for r in _read_csv(dp):
            rut_norm = _norm_rut(r.get("RUT") or "")
            mail = str(r.get("Mail") or "").strip()
            if rut_norm and mail:
                personal_mail[rut_norm] = mail

    # Cargar celular desde Numeros.xlsx → hoja "BASE DE DATOS OFICIAL"
    personal_cel: dict[str, str] = {}
    nx = _numeros_path()
    if nx.exists() and _OPENPYXL:
        wb = _openpyxl.load_workbook(nx, data_only=True, read_only=True)
        if "BASE DE DATOS OFICIAL" in wb.sheetnames:
            ws = wb["BASE DE DATOS OFICIAL"]
            rows_xl = ws.iter_rows(min_row=3, values_only=True)
            headers_xl = [str(c or "").strip() for c in next(rows_xl)]
            try:
                idx_rut = headers_xl.index("RUT")
                idx_cel = headers_xl.index("CELULAR")
            except ValueError:
                idx_rut = idx_cel = -1
            if idx_rut >= 0 and idx_cel >= 0:
                for row in rows_xl:
                    rut_norm = _norm_rut(row[idx_rut] if idx_rut < len(row) else "")
                    cel = str(row[idx_cel] if idx_cel < len(row) else "").strip()
                    if cel and cel not in ("None", "0"):
                        personal_cel[rut_norm] = cel
        wb.close()

    rows     = _read_csv(path)
    clientes: dict[str, dict] = {}

    for r in rows:
        temp = str(r.get("Temporada") or "").strip()
        if not temp.isdigit() or int(temp) < 40:
            continue

        rut      = str(r.get("Rut")          or "").strip()
        nombre   = str(r.get("cliente")       or "").strip() or rut
        ciudad   = str(r.get("Ciudad")        or "").strip()
        subcateg = str(r.get("SubCategoria")  or "").strip()
        total    = _to_int(r.get("Total"))
        cant     = _to_int(r.get("Cant"))
        num      = str(r.get("Numero")        or "").strip()
        key      = rut or nombre

        if key not in clientes:
            clientes[key] = {
                "rut":     rut,
                "nombre":  nombre,
                "ciudad":  ciudad,
                "coles":   {c: {"total": 0, "prendas": 0, "docs": set()} for c in COLECCIONES},
                "subs":    {},
                "total":   0,
                "prendas": 0,
            }

        c = clientes[key]
        if ciudad and not c["ciudad"]:
            c["ciudad"] = ciudad

        if temp in COLECCIONES:
            c["coles"][temp]["total"]   += total
            c["coles"][temp]["prendas"] += cant
            c["coles"][temp]["docs"].add(num)

        c["total"]   += total
        c["prendas"] += cant

        if subcateg:
            c["subs"][subcateg] = c["subs"].get(subcateg, 0) + cant

    result = []
    for key, c in clientes.items():
        if c["total"] == 0:
            continue

        coles_serial = {
            cole: {
                "total":   d["total"],
                "prendas": d["prendas"],
                "docs":    len(d["docs"]),
            }
            for cole, d in c["coles"].items()
        }

        top_subs = sorted(c["subs"].items(), key=lambda x: -x[1])[:5]

        rut_norm = _norm_rut(c["rut"])
        result.append({
            "key":            key,
            "rut":            c["rut"],
            "nombre":         c["nombre"],
            "ciudad":         c["ciudad"].strip(),
            "mail":           personal_mail.get(rut_norm, ""),
            "celular":        personal_cel.get(rut_norm, ""),
            "coles":          coles_serial,
            "total":          c["total"],
            "prendas":        c["prendas"],
            "top_sub": top_subs[0][0] if top_subs else "—",
            "top_subs": [{"nombre": s, "cant": p} for s, p in top_subs],
        })

    result.sort(key=lambda x: -x["total"])
    return result


# ── Vista principal ───────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template(
        "index.html",
        mejores_clientes=report_mejores_clientes(),
        no_compraron=report_no_compraron(),
        modelo_bota=report_modelo_bota(),
        ficha_clientes=report_ficha_clientes(),
    )
